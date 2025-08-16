#!/usr/bin/env python3
"""
Excel Export Module for Trading Reports
"""

import os
import logging
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from database.service import get_db_service
from performance_analytics import PerformanceAnalytics

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Export trading data and reports to Excel"""
    
    def __init__(self):
        self.db = get_db_service()
        self.analytics = PerformanceAnalytics()
        
        # Styling
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.profit_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        self.loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_trading_report(
        self,
        user_id: int,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        output_path: str = "reports/"
    ) -> str:
        """Export comprehensive trading report to Excel"""
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add sheets
        self._create_summary_sheet(wb, user_id, start_date, end_date)
        self._create_trades_sheet(wb, user_id, start_date, end_date)
        self._create_positions_sheet(wb, user_id)
        self._create_performance_sheet(wb, user_id, start_date, end_date)
        self._create_pairs_analysis_sheet(wb, user_id, start_date, end_date)
        self._create_strategy_sheet(wb, user_id, start_date, end_date)
        self._create_risk_metrics_sheet(wb, user_id, start_date, end_date)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{output_path}trading_report_{user_id}_{timestamp}.xlsx"
        
        # Ensure directory exists
        os.makedirs(output_path, exist_ok=True)
        
        # Save workbook
        wb.save(filename)
        logger.info(f"Excel report saved to {filename}")
        
        return filename
    
    def _create_summary_sheet(self, wb: Workbook, user_id: int, start_date, end_date):
        """Create summary sheet with key metrics"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "TRADING REPORT SUMMARY"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        # Date range
        ws['A3'] = "Report Period:"
        ws['B3'] = f"{start_date.strftime('%Y-%m-%d') if start_date else 'All Time'} to {end_date.strftime('%Y-%m-%d') if end_date else 'Present'}"
        
        # Get metrics
        trades = self.db.get_user_trades(user_id, start_date, end_date)
        metrics = self.analytics.calculate_metrics(trades)
        
        # Key metrics
        row = 5
        metrics_data = [
            ("Total Trades", len(trades)),
            ("Win Rate", f"{metrics.get('win_rate', 0):.2f}%"),
            ("Total P&L", f"${metrics.get('total_pnl', 0):,.2f}"),
            ("Average Trade", f"${metrics.get('avg_trade', 0):.2f}"),
            ("Best Trade", f"${metrics.get('best_trade', 0):.2f}"),
            ("Worst Trade", f"${metrics.get('worst_trade', 0):.2f}"),
            ("Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}"),
            ("Max Drawdown", f"{metrics.get('max_drawdown', 0):.2f}%"),
            ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}"),
            ("Recovery Factor", f"{metrics.get('recovery_factor', 0):.2f}")
        ]
        
        for label, value in metrics_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Style columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        # Add chart
        self._add_equity_chart(ws, trades, start_row=5, start_col=4)
    
    def _create_trades_sheet(self, wb: Workbook, user_id: int, start_date, end_date):
        """Create detailed trades sheet"""
        ws = wb.create_sheet("Trades")
        
        # Get trades
        trades = self.db.get_user_trades(user_id, start_date, end_date)
        
        # Convert to DataFrame
        df = pd.DataFrame(trades)
        
        if not df.empty:
            # Add calculated columns
            df['Duration'] = (df['exit_time'] - df['entry_time']).dt.total_seconds() / 3600  # Hours
            df['ROI %'] = (df['pnl'] / df['position_size']) * 100
            
            # Write headers
            headers = ['ID', 'Symbol', 'Side', 'Entry Time', 'Entry Price', 
                      'Exit Time', 'Exit Price', 'Size', 'P&L', 'ROI %', 'Duration (hrs)']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, row_data in df.iterrows():
                row_num = row_idx + 2
                
                ws.cell(row=row_num, column=1, value=row_data.get('id'))
                ws.cell(row=row_num, column=2, value=row_data.get('symbol'))
                ws.cell(row=row_num, column=3, value=row_data.get('side'))
                ws.cell(row=row_num, column=4, value=row_data.get('entry_time'))
                ws.cell(row=row_num, column=5, value=row_data.get('entry_price'))
                ws.cell(row=row_num, column=6, value=row_data.get('exit_time'))
                ws.cell(row=row_num, column=7, value=row_data.get('exit_price'))
                ws.cell(row=row_num, column=8, value=row_data.get('position_size'))
                
                # P&L with color
                pnl_cell = ws.cell(row=row_num, column=9, value=row_data.get('pnl'))
                if row_data.get('pnl', 0) > 0:
                    pnl_cell.fill = self.profit_fill
                else:
                    pnl_cell.fill = self.loss_fill
                
                ws.cell(row=row_num, column=10, value=row_data.get('ROI %'))
                ws.cell(row=row_num, column=11, value=row_data.get('Duration'))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_positions_sheet(self, wb: Workbook, user_id: int):
        """Create current positions sheet"""
        ws = wb.create_sheet("Current Positions")
        
        # Get positions
        positions = self.db.get_user_positions(user_id)
        
        if positions:
            # Headers
            headers = ['Symbol', 'Side', 'Size', 'Entry Price', 'Current Price', 
                      'Unrealized P&L', 'ROI %', 'Duration', 'Leverage']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # Data
            for row_idx, pos in enumerate(positions, 2):
                ws.cell(row=row_idx, column=1, value=pos.get('symbol'))
                ws.cell(row=row_idx, column=2, value=pos.get('side'))
                ws.cell(row=row_idx, column=3, value=pos.get('size'))
                ws.cell(row=row_idx, column=4, value=pos.get('entry_price'))
                ws.cell(row=row_idx, column=5, value=pos.get('current_price'))
                
                # Unrealized P&L with color
                pnl_cell = ws.cell(row=row_idx, column=6, value=pos.get('unrealized_pnl'))
                if pos.get('unrealized_pnl', 0) > 0:
                    pnl_cell.fill = self.profit_fill
                else:
                    pnl_cell.fill = self.loss_fill
                
                ws.cell(row=row_idx, column=7, value=pos.get('roi_percent'))
                ws.cell(row=row_idx, column=8, value=pos.get('duration_hours'))
                ws.cell(row=row_idx, column=9, value=pos.get('leverage'))
    
    def _create_performance_sheet(self, wb: Workbook, user_id: int, start_date, end_date):
        """Create performance metrics sheet"""
        ws = wb.create_sheet("Performance")
        
        # Monthly performance
        ws['A1'] = "MONTHLY PERFORMANCE"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Get monthly data
        monthly_data = self.analytics.get_monthly_performance(user_id, start_date, end_date)
        
        if monthly_data:
            # Headers
            headers = ['Month', 'Trades', 'Win Rate %', 'Total P&L', 'Avg Trade', 'Best', 'Worst']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # Data
            for row_idx, month_data in enumerate(monthly_data, 4):
                ws.cell(row=row_idx, column=1, value=month_data['month'])
                ws.cell(row=row_idx, column=2, value=month_data['trades'])
                ws.cell(row=row_idx, column=3, value=f"{month_data['win_rate']:.1f}")
                ws.cell(row=row_idx, column=4, value=f"${month_data['total_pnl']:,.2f}")
                ws.cell(row=row_idx, column=5, value=f"${month_data['avg_trade']:.2f}")
                ws.cell(row=row_idx, column=6, value=f"${month_data['best_trade']:.2f}")
                ws.cell(row=row_idx, column=7, value=f"${month_data['worst_trade']:.2f}")
            
            # Add chart
            self._add_monthly_chart(ws, monthly_data, start_row=3, start_col=9)
    
    def _create_pairs_analysis_sheet(self, wb: Workbook, user_id: int, start_date, end_date):
        """Create analysis by trading pairs"""
        ws = wb.create_sheet("Pairs Analysis")
        
        # Get pair statistics
        pair_stats = self.analytics.get_pair_statistics(user_id, start_date, end_date)
        
        if pair_stats:
            # Headers
            headers = ['Pair', 'Trades', 'Win Rate %', 'Total P&L', 'Avg P&L', 
                      'Best Trade', 'Worst Trade', 'Total Volume']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # Sort by total P&L
            pair_stats.sort(key=lambda x: x['total_pnl'], reverse=True)
            
            # Data
            for row_idx, pair_data in enumerate(pair_stats, 2):
                ws.cell(row=row_idx, column=1, value=pair_data['symbol'])
                ws.cell(row=row_idx, column=2, value=pair_data['trades'])
                ws.cell(row=row_idx, column=3, value=f"{pair_data['win_rate']:.1f}")
                
                # P&L with color
                pnl_cell = ws.cell(row=row_idx, column=4, value=f"${pair_data['total_pnl']:,.2f}")
                if pair_data['total_pnl'] > 0:
                    pnl_cell.fill = self.profit_fill
                else:
                    pnl_cell.fill = self.loss_fill
                
                ws.cell(row=row_idx, column=5, value=f"${pair_data['avg_pnl']:.2f}")
                ws.cell(row=row_idx, column=6, value=f"${pair_data['best_trade']:.2f}")
                ws.cell(row=row_idx, column=7, value=f"${pair_data['worst_trade']:.2f}")
                ws.cell(row=row_idx, column=8, value=f"${pair_data['total_volume']:,.2f}")
            
            # Add pie chart
            self._add_pairs_pie_chart(ws, pair_stats, start_row=2, start_col=10)
    
    def _create_strategy_sheet(self, wb: Workbook, user_id: int, start_date, end_date):
        """Create strategy performance sheet"""
        ws = wb.create_sheet("Strategy Performance")
        
        # Get strategy data
        strategy_data = self.analytics.get_strategy_performance(user_id, start_date, end_date)
        
        if strategy_data:
            # Headers
            headers = ['Strategy', 'Trades', 'Win Rate %', 'Total P&L', 'Sharpe Ratio', 
                      'Max Drawdown %', 'Profit Factor']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            
            # Data
            for row_idx, strat in enumerate(strategy_data, 2):
                ws.cell(row=row_idx, column=1, value=strat['name'])
                ws.cell(row=row_idx, column=2, value=strat['trades'])
                ws.cell(row=row_idx, column=3, value=f"{strat['win_rate']:.1f}")
                ws.cell(row=row_idx, column=4, value=f"${strat['total_pnl']:,.2f}")
                ws.cell(row=row_idx, column=5, value=f"{strat['sharpe_ratio']:.2f}")
                ws.cell(row=row_idx, column=6, value=f"{strat['max_drawdown']:.2f}")
                ws.cell(row=row_idx, column=7, value=f"{strat['profit_factor']:.2f}")
    
    def _create_risk_metrics_sheet(self, wb: Workbook, user_id: int, start_date, end_date):
        """Create risk metrics sheet"""
        ws = wb.create_sheet("Risk Metrics")
        
        # Calculate risk metrics
        risk_metrics = self.analytics.calculate_risk_metrics(user_id, start_date, end_date)
        
        # Risk metrics
        ws['A1'] = "RISK METRICS"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        risk_data = [
            ("Value at Risk (95%)", f"${risk_metrics.get('var_95', 0):,.2f}"),
            ("Conditional VaR (95%)", f"${risk_metrics.get('cvar_95', 0):,.2f}"),
            ("Maximum Drawdown", f"{risk_metrics.get('max_drawdown', 0):.2f}%"),
            ("Drawdown Duration", f"{risk_metrics.get('max_dd_duration', 0)} days"),
            ("Recovery Factor", f"{risk_metrics.get('recovery_factor', 0):.2f}"),
            ("Ulcer Index", f"{risk_metrics.get('ulcer_index', 0):.2f}"),
            ("Calmar Ratio", f"{risk_metrics.get('calmar_ratio', 0):.2f}"),
            ("Sortino Ratio", f"{risk_metrics.get('sortino_ratio', 0):.2f}"),
            ("Downside Deviation", f"{risk_metrics.get('downside_deviation', 0):.2f}%"),
            ("Kelly Criterion", f"{risk_metrics.get('kelly_criterion', 0):.2f}%")
        ]
        
        for label, value in risk_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Add risk distribution chart
        self._add_risk_distribution_chart(ws, risk_metrics, start_row=3, start_col=4)
    
    def _add_equity_chart(self, ws, trades, start_row, start_col):
        """Add equity curve chart"""
        if not trades:
            return
        
        # Calculate cumulative P&L
        cumulative_pnl = []
        running_total = 0
        dates = []
        
        for trade in sorted(trades, key=lambda x: x['exit_time']):
            running_total += trade['pnl']
            cumulative_pnl.append(running_total)
            dates.append(trade['exit_time'].strftime('%Y-%m-%d'))
        
        # Write data for chart
        ws.cell(row=start_row, column=start_col, value="Date")
        ws.cell(row=start_row, column=start_col + 1, value="Cumulative P&L")
        
        for i, (date, pnl) in enumerate(zip(dates, cumulative_pnl)):
            ws.cell(row=start_row + i + 1, column=start_col, value=date)
            ws.cell(row=start_row + i + 1, column=start_col + 1, value=pnl)
        
        # Create chart
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.y_axis.title = "Cumulative P&L ($)"
        chart.x_axis.title = "Date"
        
        data = Reference(ws, min_col=start_col + 1, min_row=start_row, 
                        max_col=start_col + 1, max_row=start_row + len(dates))
        categories = Reference(ws, min_col=start_col, min_row=start_row + 1,
                              max_row=start_row + len(dates))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, f"{chr(65 + start_col + 3)}{start_row}")
    
    def _add_monthly_chart(self, ws, monthly_data, start_row, start_col):
        """Add monthly performance chart"""
        if not monthly_data:
            return
        
        chart = BarChart()
        chart.title = "Monthly P&L"
        chart.y_axis.title = "P&L ($)"
        chart.x_axis.title = "Month"
        
        # Reference data
        data = Reference(ws, min_col=4, min_row=start_row,
                        max_col=4, max_row=start_row + len(monthly_data))
        categories = Reference(ws, min_col=1, min_row=start_row + 1,
                              max_row=start_row + len(monthly_data))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, f"{chr(65 + start_col)}{start_row}")
    
    def _add_pairs_pie_chart(self, ws, pair_stats, start_row, start_col):
        """Add pie chart for pair distribution"""
        if not pair_stats:
            return
        
        # Only show top 10 pairs
        top_pairs = pair_stats[:10]
        
        chart = PieChart()
        chart.title = "P&L by Trading Pair"
        
        # Reference data
        data = Reference(ws, min_col=4, min_row=start_row - 1,
                        max_col=4, max_row=start_row + len(top_pairs) - 1)
        labels = Reference(ws, min_col=1, min_row=start_row,
                          max_row=start_row + len(top_pairs) - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        ws.add_chart(chart, f"{chr(65 + start_col)}{start_row}")
    
    def _add_risk_distribution_chart(self, ws, risk_metrics, start_row, start_col):
        """Add risk distribution chart"""
        # Implementation for risk distribution visualization
        pass

# Singleton instance
_exporter = None

def get_excel_exporter() -> ExcelExporter:
    """Get singleton Excel exporter instance"""
    global _exporter
    if _exporter is None:
        _exporter = ExcelExporter()
    return _exporter

if __name__ == "__main__":
    # Test export
    exporter = get_excel_exporter()
    
    # Export report for user 1
    filename = exporter.export_trading_report(
        user_id=1,
        start_date=datetime(2024, 1, 1),
        end_date=datetime.now()
    )
    
    print(f"Report exported to: {filename}")